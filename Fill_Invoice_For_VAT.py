from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException, NoSuchWindowException
import base64
import time

def getAccounting():
    with open('accounting.txt', 'r') as txt_accounting:
        varAccount = (txt_accounting.readline()).split('|')

    return varAccount

def ShowTaxCodes():
    showTaxCodes = {
        "0106520697":{"Name": "kcoffee","CompanyName":"CÔNG TY TNHH WA COFFEE","Address":"B007,B008,B009 tháp The Manor, đường Mễ Trì, phường Mỹ Đình 1, Q.Nam Từ Liêm, Hà Nội"},
        "0107975898":{"Name": "dart","Company":"CÔNG TY CỔ PHẦN D'ART","Address":"Số 41 ngõ 40 phố Ngụy Như Kon Tum, Phường Nhân Chính, Quận Thanh Xuân, TP Hà Nội, Việt Nam"},
        "0109704124":{"Name": "trangj","Company":"CÔNG TY TNHH DỊCH VỤ VÀ THƯƠNG MẠI JH TRANG","Address":"Số nhà 9b, ngõ 39 Đình Thôn, Phường Mỹ Đình 1, Quận Nam Từ Liêm, Thành phố Hà Nội, Việt Nam"},
        "2300979534":{"Name": "sonsan","Company":"Công Ty TNHH Một Thành Viên Sơn San","Address":"211 Nguyễn Cao, Phường Ninh Xá, Thành phố Bắc Ninh, Tỉnh Bắc Ninh"}}

    for key, values in showTaxCodes.items():
        print(f"{key}")
        for k,v in values.items():
            print(f"\t{k}: {v}")

def Login(pathExcel, Workbook_Active):
    driver.get(urlLogin)

    user = driver.find_element(By.ID, "UserName")
    user.clear()
    user.send_keys(Usr)

    p4ss = driver.find_element(By.ID, "Password")
    p4ss.clear()
    p4ss.send_keys(base64.b64decode(p4ssb64).decode())
    # p4ss.send_keys(p4ssb64)

    captch = driver.find_element(By.ID, "captch")
    captch.clear()
    captch.send_keys()

    try:
        WebDriverWait(driver, 60).until(EC.title_contains("Hóa đơn điện tử"))
    except TimeoutError as e:
        pass
    
    try:
        driver.get(urlInvoice)
        WebDriverWait(driver, 10).until(EC.title_contains("Tạo mới hóa đơn"))
    except TimeoutError as e:
        driver.quit()

def FillItemsNewInvoice():
    for itemInvoiceParrent in range(1, count_row):
        if itemInvoiceParrent >= 90:
            time.sleep(2)
        elif itemInvoiceParrent >= 80:
            time.sleep(1.5)
        elif itemInvoiceParrent >= 70:
            time.sleep(1.5)
        elif itemInvoiceParrent >= 60:
            time.sleep(1.2)
        elif itemInvoiceParrent >= 50:
            time.sleep(1)
        elif itemInvoiceParrent >= 40:
            time.sleep(0.5)
        elif itemInvoiceParrent >= 30:
            time.sleep(0.2)
        else:
            pass
        itemInvoiceParrentTemp = itemInvoiceParrent+1
        isRowExists = (driver.find_element(By.XPATH, f'''//*[@id="bodyTblProduct"]/tr[{itemInvoiceParrent}]/td[2]''')).text
        # isRowExists = f'''//*[@id="bodyTblProduct"]/tr[{itemInvoiceParrent}]/td[2]'''
        if int(isRowExists) == (itemInvoiceParrent):
            for itemInvoiceChild in range(4, 8):
                char = get_column_letter(itemInvoiceChild)
                cellValue = ws[char + str(itemInvoiceParrentTemp)].value
                if itemInvoiceChild == 6 and isinstance(cellValue, float):
                    value = str(cellValue).replace('.', ',')
                else:
                    value = cellValue
                fillCol = driver.find_element(By.XPATH, f'''//*[@id="bodyTblProduct"]/tr[{itemInvoiceParrent}]/td[{itemInvoiceChild}]/input''')
                fillCol.clear()
                fillCol.send_keys(value)
        else:
            print(f"isRowExist: {int(isRowExists)} != itemInvoiceParrent: {itemInvoiceParrent}")
                
    wb.close()

"""
def ShowSignOption():
    Sign = {
        1: 'C22TBT', 
        2: 'C23TBT'
    }
    print(">> Chọn ký hiệu hóa đơn <<")
    for item in Sign:
        print(f"{item}: {Sign[item]}")
"""
def SelectSignOptionNewInvoice():
    signOption = 2
    try:
        xpathSign = f'''//*[@id="Serial"]/option[{signOption}]'''
        driver.find_element(By.XPATH, xpathSign).click()
    except TimeoutError:
        driver.quit()

"""
def ShowPaymentOption():
    Payment = {
        1: 'Thanh toán tiền mặt',
        2: 'Thanh toán chuyển khoản',
        3: 'Thanh toán thẻ tín dụng',
        4: 'Hình thức HDDT',
        5: 'Hình thức thanh toán tiền mặt hoặc chuyển khoản',
        6: 'Thanh toán bù trừ'
    }

    # Here is input but I ignored    
    print("--Chọn hình thức thanh toán--") # Select Payment Option
    for item in Payment:
        print(f"{item}: {Payment[item]}")
"""
def SelectPaymentOptionNewInvoice():
    paymentOption = 5
    try:
        xpathPayment = f'''//*[@id="PaymentMethod"]/option[{paymentOption+1}]'''
        driver.find_element(By.XPATH, xpathPayment).click()
        # WebDriverWait(driver, 10).until(EC.element_to_be_selected(xpathPayment))
    except TimeoutError:
        driver.quit()

def SetTaxCode():
    # Fill Tax Code
    taxcodeNumber = driver.find_element(By.ID, "CusTaxCode")
    taxcodeNumber.clear()
    taxcodeNumber.send_keys(taxcode)
    
    # Alert after fill Tax Code
    driver.find_element(By.ID, "GetInforTax").click()
    WebDriverWait(driver, 10).until(EC.alert_is_present())
    driver.switch_to.alert.accept()

if "__main__":
    Usr, p4ssb64 = getAccounting()
    ShowTaxCodes()
    taxcode = input("Enter a Tax Code: ")
    
    # ========== Excel ==========
    pathExcel = "Hoadonrau.xlsx"
    Workbook_Active = "Sheet1"
    
    # ========== Browser ==========
    # url login form
    urlLogin = "https://8667756621-001-tt78cadmin.vnpt-invoice.com.vn/"
    # url create new Invoice
    urlInvoice = f"https://8667756621-001-tt78cadmin.vnpt-invoice.com.vn/EInvoice/create?Pattern="+"2/001"
    
    # Excel Process
    wb = load_workbook(pathExcel, data_only=True)
    ws = wb[Workbook_Active]
    
    count_row = ws.max_row
    count_column = ws.max_column
    print(f"Max Row: {count_row} | Max Column: {count_column}")
    time.sleep(1)
    try:
        driver = webdriver.Chrome()
        Login(pathExcel, Workbook_Active)
        FillItemsNewInvoice()
        SetTaxCode()
        SelectSignOptionNewInvoice()
        SelectPaymentOptionNewInvoice()
        time.sleep(6000)
    except TimeoutError as te:
        print(te)
    except NoSuchElementException as exc:
        print(exc)
    except NoSuchWindowException:
        print("Target window already closed")